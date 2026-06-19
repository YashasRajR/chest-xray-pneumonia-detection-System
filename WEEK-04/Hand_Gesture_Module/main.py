import os
import json
import time
from datetime import datetime

import cv2
import mediapipe as mp

from mediapipe.tasks import python
from mediapipe.tasks.python import vision

from openpyxl import Workbook, load_workbook

# =====================================================
# LOAD GESTURE RECOGNIZER MODEL
# =====================================================

model_path = "gesture_recognizer.task"

base_options = python.BaseOptions(
    model_asset_path=model_path
)
options = vision.GestureRecognizerOptions(
    base_options=base_options,
    num_hands=2
)
recognizer = vision.GestureRecognizer.create_from_options(
    options
)

# =====================================================
# WEBCAM
# =====================================================

cap = cv2.VideoCapture(0)
frame_number = 0
prev_time = 0
print("Press ESC to Exit")

# =====================================================
# MAIN LOOP
# =====================================================
excel_file = "Hand_Gesture_Output.xlsx"

if not os.path.exists(excel_file):

    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Hand_Data"

    ws1.append([
        "Timestamp",
        "Frame_Number",
        "Hand_Detected",
        "Hand_Count",
        "Hand_Type",
        "Tracking_ID",
        "Center_X",
        "Center_Y",
        "BBox_X",
        "BBox_Y",
        "BBox_Width",
        "BBox_Height"
    ])

    # Sheet 2
    ws2 = wb.create_sheet("Gesture_Data")

    ws2.append([
        "Timestamp",
        "Frame_Number",
        "Gesture_Name",
        "Gesture_Confidence"
    ])

    # Sheet 3
    ws3 = wb.create_sheet("Landmarks")

    ws3.append([
        "Timestamp",
        "Frame_Number",
        "Landmark_ID",
        "X",
        "Y",
        "Z",
        "Visibility"
    ])

    wb.save(excel_file)

while True:
    success, frame = cap.read()
    if not success:
        break
    frame_number += 1

    # =========================================
    # FPS CALCULATION
    # =========================================

    current_fps_time = time.time()
    fps = (
        1 / (current_fps_time - prev_time)
        if prev_time != 0 else 0
    )
    prev_time = current_fps_time

    # =========================================
    # MIRROR VIEW
    # =========================================

    frame = cv2.flip(frame, 1)
    rgb_frame = cv2.cvtColor(
        frame,
        cv2.COLOR_BGR2RGB
    )
    mp_image = mp.Image(
        image_format=mp.ImageFormat.SRGB,
        data=rgb_frame
    )

    # =========================================
    # GESTURE RECOGNITION
    # =========================================

    result = recognizer.recognize(mp_image)
    hand_count = len(result.hand_landmarks)

    # =========================================
    # PROCESS HANDS
    # =========================================

    for hand_id, hand_landmarks in enumerate(
        result.hand_landmarks
    ):
        h, w, _ = frame.shape
        x_list = []
        y_list = []
        landmarks = []
        # -------------------------------------
        # LANDMARKS
        # -------------------------------------
        for idx, lm in enumerate(hand_landmarks):
            x = int(lm.x * w)
            y = int(lm.y * h)
            x_list.append(x)
            y_list.append(y)
            landmarks.append({
                "landmark_id": idx,
                "x": float(lm.x),
                "y": float(lm.y),
                "z": float(lm.z),
                "visibility": 1.0
            })
            cv2.circle(
                frame,
                (x, y),
                4,
                (0, 0, 255),
                -1
            )

        # -------------------------------------
        # BOUNDING BOX
        # -------------------------------------

        xmin = min(x_list)
        xmax = max(x_list)

        ymin = min(y_list)
        ymax = max(y_list)

        bbox_width = xmax - xmin
        bbox_height = ymax - ymin

        center_x = (xmin + xmax) // 2
        center_y = (ymin + ymax) // 2

        cv2.rectangle(
            frame,
            (xmin, ymin),
            (xmax, ymax),
            (255, 0, 0),
            2
        )

        # -------------------------------------
        # GESTURE INFO
        # -------------------------------------

        gesture_name = "Unknown"
        gesture_confidence = 0.0

        if hand_id < len(result.gestures):
            if len(result.gestures[hand_id]) > 0:
                top_gesture = result.gestures[hand_id][0]
                gesture_name = (
                    top_gesture.category_name
                )
                gesture_confidence = float(
                    top_gesture.score
                )

        # -------------------------------------
        # LEFT / RIGHT HAND
        # -------------------------------------

        hand_type = "Unknown"
        hand_type = "Unknown"
        if hand_id < len(result.handedness):
            detected_hand = (
                result.handedness[hand_id][0]
                .category_name
            )
            # Swap labels because webcam is mirrored
            if detected_hand == "Left":
                hand_type = "Right"
            else:
                hand_type = "Left"

        # -------------------------------------
        # DISPLAY LABELS
        # -------------------------------------

        cv2.putText(
            frame,
            hand_type,
            (xmin, ymin - 10),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            (0, 255, 0),
            2
        )
        cv2.putText(
            frame,
            gesture_name,
            (xmin, ymax + 30),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            (0, 255, 255),
            2
        )
        cv2.putText(
            frame,
            f"{gesture_confidence * 100:.1f}%",
            (xmin, ymax + 55),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.7,
            (255, 255, 255),
            2
        )

        # -------------------------------------
        # JSON OUTPUT
        # -------------------------------------

        output = {

            "session_id": "S001",
            "camera_id": "CAM01",
            "person_id": f"P{hand_id+1}",
            "timestamp": datetime.now().isoformat(),
            "frame_number": frame_number,
            "hand_data": {
                "hand_detected": True,
                "hand_count": hand_count,
                "hand_type": hand_type,
                "hand_tracking_id": f"H{hand_id+1}",
                "hand_center_x": center_x,
                "hand_center_y": center_y,
                "hand_bbox_x": xmin,
                "hand_bbox_y": ymin,
                "hand_bbox_width": bbox_width,
                "hand_bbox_height": bbox_height
            },
            "gesture_data": {
                "gesture_name": gesture_name,
                "gesture_confidence": round( gesture_confidence,4)
            },
            "landmarks": landmarks
        }

        # Uncomment if you want JSON output
        # print(json.dumps(output))

        # =====================================
        # SAVE TO EXCEL EVERY 10TH FRAME
        # =====================================

        if frame_number % 10 == 0:

            wb = load_workbook(excel_file)

            # Hand Data
            ws1 = wb["Hand_Data"]

            ws1.append([
                datetime.now().isoformat(),
                frame_number,
                True,
                hand_count,
                hand_type,
                f"H{hand_id+1}",
                center_x,
                center_y,
                xmin,
                ymin,
                bbox_width,
                bbox_height
            ])

            # Gesture Data
            ws2 = wb["Gesture_Data"]

            ws2.append([
                datetime.now().isoformat(),
                frame_number,
                gesture_name,
                round(gesture_confidence, 4)
            ])

            # Landmark Data
            ws3 = wb["Landmarks"]

            for lm in landmarks:

                ws3.append([
                    datetime.now().isoformat(),
                    frame_number,
                    lm["landmark_id"],
                    lm["x"],
                    lm["y"],
                    lm["z"],
                    lm["visibility"]
                ])

            wb.save(excel_file)

    
    # =================================================
    # TOP DASHBOARD
    # =================================================

    current_time = datetime.now().strftime(
        "%H:%M:%S"
    )
    cv2.rectangle(
        frame,
        (0, 0),
        (frame.shape[1], 50),
        (40, 40, 40),
        -1
    )
    cv2.putText(
        frame,
        f"Hands: {hand_count}",
        (20, 35),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.8,
        (0, 255, 0),
        2
    )
    cv2.putText(
        frame,
        f"FPS: {fps:.1f}",
        (220, 35),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.8,
        (255, 255, 0),
        2
    )
    cv2.putText(
        frame,
        current_time,
        (450, 35),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.8,
        (0, 255, 255),
        2
    )

    # =================================================
    # FULLSCREEN DISPLAY
    # =================================================

    cv2.namedWindow(
        "Hand Gesture Module",
        cv2.WINDOW_NORMAL
    )
    cv2.setWindowProperty(
        "Hand Gesture Module",
        cv2.WND_PROP_FULLSCREEN,
        cv2.WINDOW_FULLSCREEN
    )
    cv2.imshow(
        "Hand Gesture Module",
        frame
    )

    # ESC TO EXIT

    if cv2.waitKey(1) & 0xFF == 27:
        break

# =====================================================
# CLEANUP
# =====================================================

cap.release()
cv2.destroyAllWindows()